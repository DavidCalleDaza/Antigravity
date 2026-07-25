"""
Servinow API — Exports Module: Excel Builder.

Construye archivos .xlsx con apariencia de reporte corporativo ERP usando
OpenPyXL, siguiendo la plantilla corporativa optimizada para reportes de cualquier ancho.
"""

import os
from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment

from app.modules.exports.charts import add_charts_below_summary
from app.modules.exports.constants import (
    DEFAULT_LOGO_PATH,
    FOOTER_LINE_1,
    FOOTER_LINE_2,
    HEADER_TEXT_START_COL,
    SYSTEM_SLOGAN,
    SYSTEM_VERSION,
)
from app.modules.exports.styles import (
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
    DIVIDER_BORDER,
    FOOTER_FONT,
    FORMAT_CURRENCY,
    FORMAT_DATE,
    FORMAT_INTEGER,
    FORMAT_PERCENT,
    METADATA_LABEL_FONT,
    METADATA_VALUE_FONT,
    REPORT_TITLE_FONT,
    SECTION_TITLE_FILL,
    SECTION_TITLE_FONT,
    SUBTITLE_FONT,
    SUMMARY_FILL,
    SUMMARY_LABEL_FONT,
    SUMMARY_VALUE_FONT,
    TABLE_HEADER_FILL,
    TABLE_HEADER_FONT,
    THIN_BORDER,
    TITLE_FONT,
)
from app.modules.exports.templates import ColumnFormat, ReportData
from app.modules.exports.utils import (
    autosize_column_width,
    format_date_value,
    short_sheet_name,
    to_numeric_value,
)

# Se define un ancho mínimo de 11 columnas para que el diseño coincida con el ancho del gráfico (18 cm)
_MIN_HEADER_COLUMNS = 5


def _last_col_letter(num_columns: int) -> str:
    return get_column_letter(max(num_columns, _MIN_HEADER_COLUMNS))


def _draw_divider(ws, row: int, num_columns: int) -> int:
    """Dibuja la línea divisoria corporativa (borde dorado) a lo largo de todo el ancho del reporte."""
    limit_col = max(num_columns, _MIN_HEADER_COLUMNS)
    for col_idx in range(1, limit_col + 1):
        ws.cell(row=row, column=col_idx).border = DIVIDER_BORDER
    ws.row_dimensions[row].height = 4
    return row + 2


def _write_section_title(ws, row: int, title: str, num_columns: int) -> int:
    """Escribe un título de sección y lo extiende en todo el ancho del reporte."""
    last_col_letter = _last_col_letter(num_columns)
    ws.merge_cells(f"A{row}:{last_col_letter}{row}")
    cell = ws[f"A{row}"]
    cell.value = title.upper()
    cell.font = SECTION_TITLE_FONT
    cell.fill = SECTION_TITLE_FILL
    cell.alignment = ALIGN_LEFT
    return row + 2


def _insert_logo(ws, logo_path: str | None) -> bool:
    """Inserta el logo corporativo, ajustado al ancho de la columna A."""
    path = logo_path or DEFAULT_LOGO_PATH
    if not path or not os.path.isfile(path):
        return False
    try:
        from openpyxl.drawing.image import Image as XLImage

        img = XLImage(path)
        img.height = 100
        img.width = 110
        ws.add_image(img, "A1")
        return True
    except Exception:
        return False


def _write_header_block(ws, report: ReportData, num_columns: int) -> int:
    """Escribe LOGO a la izquierda y los títulos centrados en las columnas restantes."""
    num_columns = max(num_columns, _MIN_HEADER_COLUMNS)
    last_col_letter = _last_col_letter(num_columns)
    text_col_letter = get_column_letter(HEADER_TEXT_START_COL)

    # Altura de filas de encabezado
    HEADER_ROW_HEIGHT = 22
    for row_idx in (1, 2, 3):
        ws.row_dimensions[row_idx].height = HEADER_ROW_HEIGHT

    # Combinar espacio para el logo únicamente en la columna A
    ws.merge_cells("A1:A3")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.column_dimensions["A"].width = 12  # ancho suficiente para el logo de 78px sin invadir columna B
    _insert_logo(ws, report.metadata.logo_path)

    # Títulos centrados horizontalmente en el espacio restante (de la columna C hasta la última)
    ws.merge_cells(f"{text_col_letter}1:{last_col_letter}1")
    ws[f"{text_col_letter}1"] = report.metadata.system_name.upper()
    ws[f"{text_col_letter}1"].font = TITLE_FONT
    ws[f"{text_col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"{text_col_letter}2:{last_col_letter}2")
    ws[f"{text_col_letter}2"] = SYSTEM_SLOGAN
    ws[f"{text_col_letter}2"].font = SUBTITLE_FONT
    ws[f"{text_col_letter}2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(f"{text_col_letter}3:{last_col_letter}3")
    ws[f"{text_col_letter}3"] = report.metadata.report_title.upper()
    ws[f"{text_col_letter}3"].font = REPORT_TITLE_FONT
    ws[f"{text_col_letter}3"].alignment = Alignment(horizontal="center", vertical="center")

    row = _draw_divider(ws, 5, num_columns)
    row = _write_section_title(ws, row, "Información General", num_columns)

    meta_pairs = [
        ("Fecha de generación:", datetime.now().strftime("%d/%m/%Y %H:%M")),
        ("Usuario:", report.metadata.generated_by),
        ("Empresa:", report.metadata.company),
    ]
    if report.metadata.period_label:
        meta_pairs.append(("Período:", report.metadata.period_label))
    if report.metadata.filters:
        filtros_txt = " + ".join(str(v) for v in report.metadata.filters.values())
        meta_pairs.append(("Filtros:", filtros_txt))

    for label, value in meta_pairs:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = METADATA_LABEL_FONT
        ws.merge_cells(f"B{row}:{last_col_letter}{row}")
        ws[f"B{row}"] = value
        ws[f"B{row}"].font = METADATA_VALUE_FONT
        row += 1

    row += 1
    return _draw_divider(ws, row, num_columns)


def _write_table(ws, report: ReportData, start_row: int) -> int:
    """Escribe la tabla de datos con formato nativo."""
    header_row = start_row
    for col_idx, col in enumerate(report.columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col.label)
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    for row_offset, row_data in enumerate(report.rows):
        excel_row = header_row + 1 + row_offset
        for col_idx, col in enumerate(report.columns, start=1):
            raw_value = row_data.get(col.key)
            cell = ws.cell(row=excel_row, column=col_idx)

            if col.format == ColumnFormat.CURRENCY:
                cell.value = to_numeric_value(raw_value)
                cell.number_format = FORMAT_CURRENCY
                cell.alignment = ALIGN_RIGHT
            elif col.format == ColumnFormat.PERCENT:
                cell.value = to_numeric_value(raw_value)
                cell.number_format = FORMAT_PERCENT
                cell.alignment = ALIGN_RIGHT
            elif col.format == ColumnFormat.INTEGER:
                cell.value = to_numeric_value(raw_value)
                cell.number_format = FORMAT_INTEGER
                cell.alignment = ALIGN_CENTER
            elif col.format == ColumnFormat.DATE:
                cell.value = format_date_value(raw_value)
                cell.number_format = FORMAT_DATE
                cell.alignment = ALIGN_CENTER
            else:
                cell.value = raw_value
                cell.alignment = ALIGN_LEFT

            cell.border = THIN_BORDER

    last_row = header_row + len(report.rows)
    last_col_letter = get_column_letter(len(report.columns))

    if report.rows:
        table_ref = f"A{header_row}:{last_col_letter}{last_row}"
        table = Table(displayName=f"Tabla{header_row}", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)

    # Dimensionamiento automático de las columnas con datos
    COLUMN_WIDTHS = {
        "Tipo": 18,
        "Descripción": 40,
        "Cantidad": 16,
        "Ingresos": 18,
    }

    DEFAULT_EXTRA_COLUMN_WIDTH = 5  # ancho para columnas del encabezado/gráfico que no tiene la tabla

    for col_idx, col in enumerate(report.columns, start=1):
        width = COLUMN_WIDTHS.get(col.label)

        if width is None:
            width = col.width or autosize_column_width(
                [r.get(col.key) for r in report.rows],
                col.label
            )

        ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Las columnas que están más allá de la tabla (usadas solo por el
        # encabezado/línea divisoria/gráfico, hasta _MIN_HEADER_COLUMNS) también
        # necesitan un ancho fijo — si no, quedan angostas por defecto y
        # descuadran visualmente todo el bloque superior del reporte.
    for col_idx in range(len(report.columns) + 1, _MIN_HEADER_COLUMNS + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = DEFAULT_EXTRA_COLUMN_WIDTH

    row = last_row + 2 # Esto deja un espacio entre la tabla y el resumen
    return _draw_divider(ws, row, len(report.columns))


def _write_summary(ws, report: ReportData, start_row: int, num_columns: int) -> int:
    """Escribe el bloque de Resumen Ejecutivo (KPIs)."""
    if not report.summary:
        return start_row

    row = _write_section_title(ws, start_row, "Resumen Ejecutivo", num_columns)
    for item in report.summary:
        ws[f"A{row}"] = item.label.upper()
        ws[f"A{row}"].font = SUMMARY_LABEL_FONT
        ws[f"A{row}"].fill = SUMMARY_FILL

        cell = ws[f"B{row}"]
        if item.format == ColumnFormat.CURRENCY:
            cell.value = to_numeric_value(item.value)
            cell.number_format = FORMAT_CURRENCY
        elif item.format == ColumnFormat.PERCENT:
            cell.value = to_numeric_value(item.value)
            cell.number_format = FORMAT_PERCENT
        elif item.format == ColumnFormat.INTEGER:
            cell.value = to_numeric_value(item.value)
            cell.number_format = FORMAT_INTEGER
        else:
            cell.value = item.value
        cell.font = SUMMARY_VALUE_FONT
        cell.fill = SUMMARY_FILL
        row += 1

    row += 1
    return _draw_divider(ws, row, num_columns)


def _write_footer(ws, start_row: int, num_columns: int) -> None:
    """Escribe el pie de página institucional."""
    last_col_letter = _last_col_letter(num_columns)
    lines = [
        FOOTER_LINE_1,
        FOOTER_LINE_2,
        f"Versión del sistema: {SYSTEM_VERSION}",
        f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
    ]
    row = start_row
    for line in lines:
        ws.merge_cells(f"A{row}:{last_col_letter}{row}")
        cell = ws[f"A{row}"]
        cell.value = line
        cell.font = FOOTER_FONT
        cell.alignment = ALIGN_LEFT
        row += 1


def build_excel_report(report: ReportData) -> BytesIO:
    """Construye el reporte Excel completo sin cuadrícula y con ancho alineado al gráfico."""
    wb = Workbook()
    ws = wb.active
    ws.title = short_sheet_name(report.metadata.sheet_name or report.metadata.report_title)

    # Fondo blanco (sin líneas de cuadrícula)
    ws.sheet_view.showGridLines = False

    num_columns = len(report.columns)

    next_row = _write_header_block(ws, report, num_columns)          # LOGO a la izq. + títulos centrados
    table_header_row = next_row
    next_row = _write_table(ws, report, next_row)                    # Tabla sin congelar paneles
    next_row = _write_summary(ws, report, next_row, num_columns)     # Resumen
    next_row = add_charts_below_summary(                             # Gráfico centrado abajo
        ws, report, table_header_row=table_header_row, start_row=next_row
    )
    _write_footer(ws, next_row, num_columns)                         # Pie de página extendido

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer