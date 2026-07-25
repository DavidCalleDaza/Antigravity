"""
Servinow API — Exports Module: Styles.

Paleta corporativa y estilos reutilizables de OpenPyXL, independientes
de cualquier módulo. Cambiar la identidad visual del reporte se hace
únicamente aquí.
"""

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Paleta corporativa Servinow ("Royal Velvet" — dorado + fondo oscuro)
COLOR_PRIMARY = "1C1924"      # fondo de encabezado corporativo
COLOR_GOLD = "D4AF37"         # acento dorado
COLOR_HEADER_TEXT = "FFFFFF"
COLOR_ROW_ALT = "F4F4F5"
COLOR_BORDER = "D9D9D9"
COLOR_MUTED_TEXT = "666666"

# --- Encabezado / portada ---
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=COLOR_GOLD)
SUBTITLE_FONT = Font(name="Calibri", size=11, color=COLOR_MUTED_TEXT)
REPORT_TITLE_FONT = Font(name="Calibri", size=13, bold=True, color=COLOR_PRIMARY)

# --- Metadata (Información General) ---
METADATA_LABEL_FONT = Font(name="Calibri", size=9, bold=True, color="888888")
METADATA_VALUE_FONT = Font(name="Calibri", size=10, color=COLOR_PRIMARY)

# --- Títulos de sección ("Información General", "Resumen Ejecutivo") ---
SECTION_TITLE_FONT = Font(name="Calibri", size=11, bold=True, color=COLOR_PRIMARY)
SECTION_TITLE_FILL = PatternFill(start_color=COLOR_ROW_ALT, end_color=COLOR_ROW_ALT, fill_type="solid")

# --- Línea divisoria corporativa entre secciones ---
DIVIDER_BORDER = Border(bottom=Side(style="medium", color=COLOR_GOLD))

# --- Tabla ---
TABLE_HEADER_FONT = Font(name="Calibri", size=10, bold=True, color=COLOR_HEADER_TEXT)
TABLE_HEADER_FILL = PatternFill(start_color=COLOR_PRIMARY, end_color=COLOR_PRIMARY, fill_type="solid")
TABLE_ROW_ALT_FILL = PatternFill(start_color=COLOR_ROW_ALT, end_color=COLOR_ROW_ALT, fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color=COLOR_BORDER),
    right=Side(style="thin", color=COLOR_BORDER),
    top=Side(style="thin", color=COLOR_BORDER),
    bottom=Side(style="thin", color=COLOR_BORDER),
)

# --- Resumen Ejecutivo ---
SUMMARY_LABEL_FONT = Font(name="Calibri", size=10, bold=True, color=COLOR_PRIMARY)
SUMMARY_VALUE_FONT = Font(name="Calibri", size=12, bold=True, color=COLOR_GOLD)
SUMMARY_FILL = PatternFill(start_color="FAF6EC", end_color="FAF6EC", fill_type="solid")

# --- Pie de página ---
FOOTER_FONT = Font(name="Calibri", size=8, italic=True, color="999999")

# --- Alineación ---
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)

# --- Formatos numéricos nativos de Excel ---
FORMAT_CURRENCY = '"$"#,##0'
FORMAT_PERCENT = '0.00"%"'
FORMAT_DATE = "DD/MM/YYYY"
FORMAT_INTEGER = "#,##0"

# --- Paleta de gráficos (misma identidad corporativa, sin colores genéricos de Excel) ---
CHART_COLORS = ["1C1924", "D4AF37", "8C7B4A", "4A4458", "B08D57", "6E6658"]