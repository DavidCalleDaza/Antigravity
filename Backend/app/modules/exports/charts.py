"""
DonApp API — Exports Module: Charts.

Genera gráficos nativos de Excel (NUNCA imágenes) a partir de ChartSpec,
respetando las reglas de diseño del reporte corporativo:
  - título visible
  - sin leyenda cuando hay una única serie
  - colores corporativos distintos por barra/segmento (nunca la paleta genérica de Excel)
  - tamaño proporcional a la hoja
  - ubicación SIEMPRE debajo del Resumen Ejecutivo, nunca al lado de la tabla
  - si el reporte tiene muchas filas, el gráfico se limita a un Top N
    (ordenado por value_column) para seguir siendo legible — la TABLA
    siempre muestra el 100% de los datos, solo el gráfico se recorta.
"""

from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.worksheet.worksheet import Worksheet

from app.modules.exports.constants import CHART_HEIGHT_CM, CHART_STACK_ROW_HEIGHT, CHART_WIDTH_CM
from app.modules.exports.styles import CHART_COLORS
from app.modules.exports.templates import ChartSpec, ReportData
from app.modules.exports.utils import to_numeric_value

# Columna donde se escribe el bloque de datos auxiliar (oculto) para
# gráficos con top_n. Suficientemente lejos de la tabla/encabezado
# visibles para no chocar con ninguna columna real del reporte.
_SCRATCH_COLUMN = 20  # columna T


def _write_scratch_data(ws: Worksheet, spec: ChartSpec, report: ReportData, anchor_row: int) -> tuple[Reference, Reference, int]:
    """
    Escribe un bloque de datos auxiliar y OCULTO con el Top N de filas
    (ordenadas desc. por value_column), para que el gráfico grafique solo
    ese subconjunto sin alterar el orden ni el contenido de la tabla visible.
    Retorna (data_ref, cats_ref, filas_usadas).
    """
    sorted_rows = sorted(
        report.rows,
        key=lambda r: to_numeric_value(r.get(spec.value_column)) or 0,
        reverse=True,
    )
    top_rows = sorted_rows[: spec.top_n]

    cat_col = _SCRATCH_COLUMN
    val_col = _SCRATCH_COLUMN + 1

    ws.cell(row=anchor_row, column=cat_col, value=spec.category_column)
    ws.cell(row=anchor_row, column=val_col, value=spec.value_column)

    for i, row_data in enumerate(top_rows, start=1):
        ws.cell(row=anchor_row + i, column=cat_col, value=row_data.get(spec.category_column))
        ws.cell(row=anchor_row + i, column=val_col, value=to_numeric_value(row_data.get(spec.value_column)))

    last_row = anchor_row + len(top_rows)

    from openpyxl.utils import get_column_letter
    ws.column_dimensions[get_column_letter(cat_col)].hidden = True
    ws.column_dimensions[get_column_letter(val_col)].hidden = True

    data_ref = Reference(ws, min_col=val_col, min_row=anchor_row, max_row=last_row)
    cats_ref = Reference(ws, min_col=cat_col, min_row=anchor_row + 1, max_row=last_row)
    return data_ref, cats_ref, len(top_rows)


def _build_single_chart(ws: Worksheet, spec: ChartSpec, report: ReportData, table_header_row: int, scratch_row: int):
    """Construye un único gráfico nativo a partir de un ChartSpec. Retorna None si las columnas no existen."""
    col_keys = [col.key for col in report.columns]
    try:
        cat_idx = col_keys.index(spec.category_column) + 1
        val_idx = col_keys.index(spec.value_column) + 1
    except ValueError:
        return None  # columna referenciada no existe en este reporte: se omite el gráfico, no se rompe el archivo

    if spec.top_n and len(report.rows) > spec.top_n:
        # Muchas filas: graficar solo el Top N vía bloque de datos auxiliar oculto.
        data_ref, cats_ref, plotted_rows = _write_scratch_data(ws, spec, report, scratch_row)
        chart_title = f"{spec.title} (Top {plotted_rows})"
    else:
        # Pocas filas: usar directamente el rango de la tabla visible, sin recorte.
        last_data_row = table_header_row + len(report.rows)
        data_ref = Reference(ws, min_col=val_idx, min_row=table_header_row, max_row=last_data_row)
        cats_ref = Reference(ws, min_col=cat_idx, min_row=table_header_row + 1, max_row=last_data_row)
        chart_title = spec.title

    if spec.chart_type == "pie":
        chart = PieChart()
        chart.varyColors = True  # cada segmento con su propio color, automático
    elif spec.chart_type == "line":
        chart = LineChart()
    else:
        # Columnas verticales por defecto — se leen mejor que barras horizontales
        # cuando hay pocas categorías, y evitan el efecto "bloque apilado".
        chart = BarChart()
        chart.type = "col"
        chart.gapWidth = 60    # espacio entre grupos de categorías (default 150 = demasiado compacto)
        chart.overlap = -10    # separación leve entre series si hubiera más de una

    chart.title = chart_title
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = CHART_WIDTH_CM
    chart.height = CHART_HEIGHT_CM
    chart.style = None  # evita el estilo genérico por defecto de OpenPyXL

    # Regla del prompt: eliminar leyenda cuando existe una única serie (no aporta información)
    if len(chart.series) <= 1:
        chart.legend = None

    # Colores: con una sola serie de barras, se colorea CADA BARRA (punto de
    # dato) con un color distinto de la paleta corporativa. Con varias series
    # (o líneas), cada serie mantiene su propio color, como antes.
    if len(chart.series) == 1 and spec.chart_type == "bar":
        series = chart.series[0]
        num_points = data_ref.max_row - data_ref.min_row  # excluye la fila de título
        series.data_points = [
            DataPoint(
                idx=i,
                spPr=GraphicalProperties(
                    solidFill=CHART_COLORS[i % len(CHART_COLORS)],
                    ln=LineProperties(noFill=True),
                ),
            )
            for i in range(num_points)
        ]
    else:
        for i, series in enumerate(chart.series):
            color = CHART_COLORS[i % len(CHART_COLORS)]
            series.graphicalProperties.solidFill = color
            if spec.chart_type != "pie":
                series.graphicalProperties.line.noFill = True
            if spec.chart_type == "line":
                series.graphicalProperties.line.solidFill = color
                series.graphicalProperties.line.width = 20000  # EMU, línea visible

    # Ejes visibles, SIN título de eje separado (Excel/OpenPyXL no reservan
    # espacio propio y termina superponiéndolo sobre las etiquetas).
    if spec.chart_type != "pie":
        chart.x_axis.delete = False
        chart.y_axis.delete = False
        chart.y_axis.majorGridlines.spPr = GraphicalProperties(
            ln=LineProperties(solidFill="E0E0E0", w=6350)
        )

    return chart


def add_charts_below_summary(ws: Worksheet, report: ReportData, table_header_row: int, start_row: int) -> int:
    """
    Inserta todos los gráficos del reporte, apilados verticalmente, comenzando en
    `start_row` — que siempre debe ser la fila libre DESPUÉS del Resumen Ejecutivo.
    Nunca se ancla al lado de la tabla, cumpliendo el orden: Tabla -> Resumen -> Gráfico.
    Retorna la fila siguiente libre tras el último gráfico insertado.
    """
    if not report.charts or not report.rows:
        return start_row

    anchor_row = start_row
    scratch_row = table_header_row  # los bloques auxiliares ocultos comparten el mismo nivel de fila que la tabla, en columnas lejanas
    any_added = False
    for spec in report.charts:
        chart = _build_single_chart(ws, spec, report, table_header_row, scratch_row)
        if chart is None:
            continue
        ws.add_chart(chart, f"A{anchor_row}")
        anchor_row += CHART_STACK_ROW_HEIGHT
        any_added = True

    return anchor_row + 1 if any_added else start_row